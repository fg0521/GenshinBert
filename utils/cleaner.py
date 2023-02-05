import json
import re
import pandas as pd
from openpyxl import Workbook,load_workbook
# from openpyxl.reader.excel import load_workbook
import os

def clear(x):
    x = re.sub('ã€ã€Œ', 'ã€ã€ã€Œ', x)
    x = re.sub('\n|\t|\r|Â |â€‹|â€»|ã€“|ï»¿|â—|\d{8}|={1,}|â€¦+|â–Œ|ã€|ã€Ž|ã€|ã€Œ|>>(.*?)<<|â€|â€œ|â€¢ ', '', str(x))
    x = re.sub('ï¼+', 'ï¼', x)
    x = re.sub('!+', '!', x)
    url_pattern = r'((?:(?<=[^a-zA-Z0-9]){0,}(?:(?:https?\:\/\/){0,1}(?:[a-zA-Z0-9\%]{1,}\:[a-zA-Z0-9\%]{1,}[@]){,1})(?:(?:\w{1,}\.{1}){1,5}(?:(?:[a-zA-Z]){1,})|(?:[a-zA-Z]{1,}\/[0-9]{1,3}\.[0-9]{1,3}\.[0-9]{1,3}\.[0-9]{1,3}\:[0-9]{1,4}){1})){1}(?:(?:(?:\/{0,1}(?:[a-zA-Z0-9\-\_\=\-]){1,})*)(?:[?][a-zA-Z0-9\=\%\&\_\-]{1,}){0,1})(?:\.(?:[a-zA-Z0-9]){0,}){0,1})'
    x = re.sub(url_pattern, '', x)
    x = re.sub('ï¼ˆ|\(|ã€(.*?)ã€‘|\)|ï¼‰', '', x)
    x = re.sub(' {2,}', ' ', x)
    x = re.sub('ã€|_|æ´»åŠ¨æ—¶é—´(.*?)59|', '', x)
    if 'CV' in x:
        x = x[:x.index('CV')]
    if '>>>' in x:
        x = x[:x.index('>>>')]
    return x
    # print(x)


def clear_notice():
    df = pd.read_csv('../dataset/notice.csv')
    df['s_content'] = df['s_content'].apply(lambda x: clear(''.join(eval(x))))
    df.drop_duplicates(subset='s_content', inplace=True)
    content = df['s_content'].tolist()
    df['len'] = df['s_content'].apply(lambda x: len(x))
    with open('../data/hmy-notice.txt', 'w') as f:
        for i in content:

            i = i.replace('!', 'ã€‚').replace('ï¼', 'ã€‚').replace('ï¼›', 'ã€‚').replace(';', 'ã€‚').replace('?', 'ï¼Ÿ')
            res = [k for k in i.split('ã€‚') if len(k) > 1]
            if 0 < len(res) < 2:
                if len(res[0]) > 30:
                    f.write(res[0] + '\n')
            else:
                for j in range(len(res) - 1):
                    s1 = res[j] if res[j].endswith('ï¼Ÿ') else res[j] + 'ã€‚'
                    s2 = res[j + 1] if res[j + 1].endswith('ï¼Ÿ') else res[j + 1] + 'ã€‚'
                    s = (s1 + s2).replace(' ', '')
                    if 256 > len(s) > 30:
                        f.write(s + '\n')
        f.close()
    # print(df['len'].value_counts())


def clear_info():
    res = []
    with open('../data/mhy-info.txt', 'r') as f:
        for line in f.readlines():
            while 'ã€ã€Œ' in line:
                idx = line.find('ã€ã€Œ')
                if re.findall('[\u4e00-\u9fa5\da-zA-Z]+', line[idx - 1]) and re.findall('[\u4e00-\u9fa5\da-zA-Z]+',
                                                                                        line[idx + 2]):
                    line = line.replace('ã€ã€Œ', 'ï¼Œ')
                else:
                    line = line.replace('ã€ã€Œ', '')
            line = re.sub('ã€|ã€Œ|ã€Ž|ã€', '', line)
            print(line)
            if len(line) > 15:
                res.append(line)
    with open('../data/mhy-info-clear.txt', 'w') as f1:
        [f1.write(i) for i in res]


def clear_audio():
    res = []
    symbol = set()
    rep_dict ={'~ï¼':'~','Â·Â·Â·ï¼Ÿ':'ï¼Ÿ','â€¦.':'...','~~':'~','&nbsp;':'','â€¦':'...',' ':'',',':'ï¼Œ','Â ':'','â–ˆâ–ˆ':'',
               'Â·Â·Â·':'...','ï¼Ÿï¼':'ï¼Ÿ','ï¼ã€‚':'ï¼'}
    with open('../dataset/mhy-audio.txt') as f:
        for line in f.readlines():
            while 'ã€ã€Œ' in line:
                idx = line.find('ã€ã€Œ')
                if re.findall('[\u4e00-\u9fa5\da-zA-Z]+', line[idx - 1]) and re.findall('[\u4e00-\u9fa5\da-zA-Z]+',
                                                                                        line[idx + 2]):
                    line = line.replace('ã€ã€Œ', 'ï¼Œ')
                else:
                    line = line.replace('ã€ã€Œ', '')
            line = re.sub('ã€|ã€Œ|ã€Ž|ã€', '', line)
            if len(re.findall('[\u4e00-\u9fa5]', line)) >= 20 and 'ç‰¹æ®Šæ–™ç†' not in line:
                for k,v in rep_dict.items():
                    line = line.replace(k,v)
                [symbol.add(j) for j in re.findall('[^\u4e00-\u9fa5\da-zA-Z]+',line)]
                res.append(line)
    with open('../data/mhy-audio-clear.txt', 'w') as f1:
        [f1.write(i) for i in res]
    print(symbol)


def clear_xlsx():
    wb = load_workbook('../dataset/åŽŸç¥žè¯­éŸ³æ–‡æœ¬.xlsx')
    sheets = wb.worksheets  # èŽ·å–å½“å‰æ‰€æœ‰çš„sheet
    print(sheets)

    with open('../data/mhy-paimeng-clear.txt','w') as f:
        for sheet in sheets:
            print(sheet)
            try:
                for col in sheet['B']:
                    if len(str(col.value))>10:
                        f.write(col.value+'\n')
            except Exception as e:
                print(e)
                continue

def clear_story():
    content = []
    texts = os.listdir('../dataset/story')
    for t in texts:
        with open(os.path.join('../dataset/story',t),'r') as f:
            res = f.read().split('\n')
        each_content = ''
        for i in res:
            each_content = each_content+i
            if len(each_content)>80 and len(each_content)<256:
                each_content = re.sub('ã€Ž|ã€|â€”{4,}|ï¼ˆ.*?ï¼‰','',each_content)
                each_content = each_content.replace('â€¦â€¦','...').replace('â€œ...â€','')
                print(len(each_content))
                if len(each_content) >30:
                    content.append(each_content)
                each_content = ''
        # print(content)
        # break
    with open('../data/hmy-story-clear.txt','w') as f1:
        [f1.write(i+'\n') for i in content]


def get_audio_text():
    with open('/Users/maoyufeng/Downloads/GenshinVoice-main/result.json') as f:
        text = json.load(f)
    # print(type(text))
    with open('../data/hmy-audio2-clear.txt','w') as f:
        for k,v in text.items():
            if v['language'] == 'CHS' and v.get('text'):
                f.write(v['text']+'\n')

def clear_audio2():
    with open('../data/hmy-audio2-clear.txt','r') as f:
        text = [i for i in f.readlines() if len(re.findall('[\u4e00-\u9fa5]',i)) > 6]
    text = list(set(text))
    [print(i) for i in text if len(i)>255]
    with open('../data/hmy-audio2-clear2.txt','w') as f:
        [f.write(i) for i in text]



def clear_hot_strategy():
    rep_dict = {
        'â‘ ':'1.','â‘¡':'2.','â‘¢':'3.','â‘£':'4.','â‘¤':'5.','â‘¥':'6.','â‘¦':'7.','â‘§':'8.','â‘¨':'9.','â‘©':'10.',
        'â€¦':'...','âž•':'+','ï¼ž':'>','ï¼‹':'+','â‰ˆ':'=',
    }
    symbol = ['ï½ž', 'Ï‰', 'â¶', 'â…¡', 'ï½€',  'â–²', 'â¸', 'â–‹',  'â‘¡', 'ï¸', 'â—†',  'ðŸ”¹', 'â‘£', '\ufeff',
     'â‘¢', 'â… ', 'â–¼','Ã—',  'â—',  'â–Œ',  'ã‚§', 'ã€', 'â·',
         'ã€‚', 'â†’', '\u200b', 'â€¢', 'â˜…',  'âœª', 'â€»',  'â˜†', 'â…¢', 'Â´', 'â†“', '\xa0', 'Â°', 'â­', 'â‘ ',  '\t', 'â‘¤', 'âˆš','>>']

    df = pd.read_csv('../dataset/hot_strategy.csv')
    df = df[['s_content']]
    df['s_content'] = df['s_content'].apply(lambda x:''.join(eval(x)).split('\n'))
    df['s_content'] = df['s_content'].apply(lambda x:[re.sub(' |_\(.*?\)','',s) for s in x if re.findall('[\u4e00-\u9fa5]+',s)])
    data = []
    for _,row in df.iterrows():
        data.extend(row['s_content'])
    new_data = []
    for s in data:
        for k,v in rep_dict.items():
            s = s.replace(k,v)
        for i in symbol:
            s = s.replace(i,'')
        if len(re.findall('[\u4e00-\u9fa5]',s))>10:
            new_data.append(s)
    # new_data = list(set(new_data))
    with open('../data/mhy-strategy-clear.txt','w') as f:
        [f.write(s+'\n') for s in new_data]
    [print(s) for s in new_data if len(s)>256]
    # symbol = set(list(''.join(re.findall('[^\da-zA-Z\u4e00-\u9fa5]',''.join(data)))))
    # print(list(symbol))


if __name__ == '__main__':
    clear_hot_strategy()
    # with open('../data/mhy-strategy-clear.txt', 'r') as f:
    #     res = [i.replace('\n','') for i in f.readlines()]
    # new_data = list(set(res))
    # with open('../data/mhy-strategy-clear.txt','w') as f:
    #     [f.write(s+'\n') for s in new_data]